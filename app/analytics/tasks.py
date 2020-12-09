from itertools import takewhile

from openpyxl import load_workbook

from perx.celery import app


@app.task
def handle_report(report_id: int) -> None:
    """Calculates result of adding or removing one positive number from file"""

    from analytics.models import Report

    report = Report.objects.get(pk=report_id)
    workbook = load_workbook(report.file)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        after_idx = before_idx = None
        for idx, (value,) in enumerate(sheet.iter_cols(min_row=1, max_row=1, values_only=True), start=1):
            if value == 'after':
                after_idx = idx
            elif value == 'before':
                before_idx = idx

            if after_idx and before_idx:
                before = (value for (value,) in
                          sheet.iter_rows(min_row=2, min_col=before_idx, max_col=before_idx, values_only=True) if value)
                after = [value for (value,) in takewhile(lambda x: x is not None,
                                                         sheet.iter_rows(min_row=2, min_col=after_idx,
                                                                         max_col=after_idx, values_only=True)) if value]

                result = report.find_diff(before=before, after=after)
                if result:
                    report.finish_processing(result)
                return
